#-*- coding: utf-8 -*-
import openpyxl
import os
gradient = ['800000', '8B0000', 'B22222','FF0000','FFD700','9ACD32', 'ADFF2F','7CFC00', '00FF00']
print('Если значения одинаковы, то ячейка будет окрашена в жёлтый, если значение увеличилось - в красный, если уменьшилось - в зелёный. В зависимости от разницы, ячейки будут окрашены в разные оттенки красного или зелёного, всего есть ' + str(int((len(gradient) - 1)/2)) +' оттенка красного и столько же зелёного')
print('Введите путь до старого файла: ')
old = input()
print('Введите путь до нового файла: ')
new = input()
path = ''
print('Введите путь для сохранения файла: ')
path = input()
from openpyxl.styles import Font, PatternFill

oldDataDil_1 = []
oldDataDil_2 = []
oldDataDil_3 = []
oldTraderPrice = []
oldClientPrice = []

newDataDil_1 = []
newDataDil_2 = []
newDataDil_3 = []
newTraderPrice = []
newClientPrice = []

wb = openpyxl.reader.excel.load_workbook(filename=old, data_only=True)
wb.active = 0
sheet1 = wb.active


for i in range(7,400):
    oldDataDil_1.append(sheet1['D' + str(i)].value)
    oldDataDil_2.append(sheet1['E' + str(i)].value)
    oldDataDil_3.append(sheet1['F' + str(i)].value)
    oldTraderPrice.append(sheet1['G' + str(i)].value)
    oldClientPrice.append(sheet1['H' + str(i)].value)


wb = openpyxl.reader.excel.load_workbook(filename=new, data_only=True)
wb.active = 0
sheet2 = wb.active

for i in range(7,400):
    newDataDil_1.append(sheet2['D' + str(i)].value)
    newDataDil_2.append(sheet2['E' + str(i)].value)
    newDataDil_3.append(sheet2['F' + str(i)].value)
    newTraderPrice.append(sheet2['G' + str(i)].value)
    newClientPrice.append(sheet2['H' + str(i)].value)

for j in range(0,352):
    if newDataDil_1[j] == None:
        newDataDil_1[j] = 0
    if newDataDil_2[j] == None:
        newDataDil_2[j] = 0
    if newDataDil_3[j] == None:
        newDataDil_3[j] = 0
    if newTraderPrice[j] == None:
        newTraderPrice[j] = 0
    if newClientPrice[j] == None:
        newClientPrice[j] = 0
    if oldDataDil_1[j] == None:
        oldDataDil_1[j] = 0
    if oldDataDil_2[j] == None:
        oldDataDil_2[j] = 0
    if oldDataDil_3[j] == None:
        oldDataDil_3[j] = 0
    if oldTraderPrice[j] == None:
        oldTraderPrice[j] = 0
    if oldClientPrice[j] == None:
        oldClientPrice[j] = 0

colorR = PatternFill(fill_type='solid', start_color='ff0000')
colorG = PatternFill(fill_type='solid', start_color='00ff00')

def color(old, new):
    if (old == None or old == 0):
        return PatternFill(fill_type = 'solid', start_color = 'FFFFFF')
    else:
        if (old - new) < 0 and abs((old - new)) > (old * 0.3):
            power = 0
        if (old - new) < 0 and abs((old - new)) > (old * 0.1):
            power = 1
        if (old - new) < 0 and abs((old - new)) > (old * 0.05):
            power = 2
        if (old - new) < 0 and abs((old - new)) <= (old * 0.05):
            power = 3
        if (old - new) == 0:
            power = 4
        if (old - new) > 0 and abs((old - new)) <= (old * 0.05):
            power = 5
        if (old - new) > 0 and abs((old - new)) > (old * 0.05):
            power = 6
        if (old - new) > 0 and abs((old - new)) > (old * 0.1):
            power = 7
        if (old - new) > 0 and abs((old - new)) > (old * 0.3):
            power = 8
    return PatternFill(fill_type = 'solid', start_color = gradient[power])

for j in range(0, sheet2.max_row - 47):
    sheet2['D' + str(j + 7)].fill = color(oldDataDil_1[j], newDataDil_1[j])
    sheet2['E' + str(j + 7)].fill = color(oldDataDil_2[j], newDataDil_2[j])
    sheet2['F' + str(j + 7)].fill = color(oldDataDil_3[j], newDataDil_3[j])
    sheet2['G' + str(j + 7)].fill = color(oldTraderPrice[j], newTraderPrice[j])
    sheet2['H' + str(j + 7)].fill = color(oldClientPrice[j], newClientPrice[j])
print('Окрашено, находится в папке: ' + path)
wb.save(path + "Окрашенный.xlsx")
input('Press ENTER to exit')
